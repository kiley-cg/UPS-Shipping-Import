#!/usr/bin/env python3
"""
UPS Tracking Import
-------------------
Reads today's UPS Excel files from Google Drive (via the Drive API),
groups shipments by Syncore PO number, and adds a Job Log entry to each
job in Syncore.

Runs as a scheduled GitHub Actions workflow — no local machine required.

Setup:
    pip install -r requirements.txt
    # Set GitHub Actions secrets (see README for the full list)
    # Trigger manually: Actions tab → UPS Tracking Import → Run workflow

Local testing:
    cp .env.example .env   # fill in all values including GOOGLE_* vars
    python ups_import.py
"""

import asyncio
import glob
import io
import json
import os
import re
import smtplib
import tempfile
from datetime import date, timedelta, timezone, datetime
from zoneinfo import ZoneInfo
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import httpx

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv optional; env vars can be set directly

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SYNCORE_USERNAME = os.environ.get("SYNCORE_USERNAME", "")
SYNCORE_PASSWORD = os.environ.get("SYNCORE_PASSWORD", "")
SYNCORE_WEB_BASE = "https://www.ateasesystems.net"

GMAIL_USER = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
CSR_EMAIL = os.environ.get("CSR_EMAIL", "")
ACCOUNTING_EMAIL = os.environ.get("ACCOUNTING_EMAIL", "accounting@colorgraphicswa.com")

# Google Drive — set via GitHub Actions secrets or .env
GOOGLE_SERVICE_ACCOUNT_JSON = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
GOOGLE_DRIVE_FOLDER_ID = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "")

# Local fallback path (only used when GOOGLE_SERVICE_ACCOUNT_JSON is not set)
LOCAL_UPS_PATH = os.environ.get(
    "UPS_DRIVE_PATH",
    "/Users/kileygustafson/Library/CloudStorage/"
    "GoogleDrive-kiley@colorgraphicswa.com/Shared drives/UPS Shipping Info",
)

HTTP_TIMEOUT = 30

# Regex: Syncore PO number  e.g. "31987-1", "32073-2"
PO_PATTERN = re.compile(r"^\d{5,6}-\d+$")

# Regex: SRF reference number  e.g. "SRF-1234", "SRF-ABC"
SRF_PATTERN = re.compile(r"^SRF-", re.IGNORECASE)

DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


# ---------------------------------------------------------------------------
# Google Drive file download
# ---------------------------------------------------------------------------

def _drive_service():
    """Build and return an authenticated Google Drive API service object."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    creds_info = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
    creds = service_account.Credentials.from_service_account_info(
        creds_info, scopes=DRIVE_SCOPES
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def download_todays_files(tmp_dir: str) -> list[str]:
    """
    Download today's UPS Excel files from Google Drive to tmp_dir.
    Returns a list of local file paths.

    Falls back to the local filesystem path if GOOGLE_SERVICE_ACCOUNT_JSON
    is not set (useful for local testing with Google Drive synced to disk).
    """
    date_prefix = datetime.now(ZoneInfo("America/Los_Angeles")).strftime("%Y%m%d")  # e.g. 20260225

    # ── Local fallback ────────────────────────────────────────────────────
    if not GOOGLE_SERVICE_ACCOUNT_JSON:
        print(f"  [Drive API not configured — using local path: {LOCAL_UPS_PATH}]")
        files: list[str] = []
        for ext in ("xlsx", "xls", "XLSX", "XLS", "csv", "CSV"):
            files.extend(glob.glob(os.path.join(LOCAL_UPS_PATH, f"{date_prefix}*.{ext}")))
        # Deduplicate
        seen: set[str] = set()
        return [f for f in files if not (f in seen or seen.add(f))]  # type: ignore[func-returns-value]

    # ── Google Drive API ──────────────────────────────────────────────────
    if not GOOGLE_DRIVE_FOLDER_ID:
        raise RuntimeError(
            "GOOGLE_DRIVE_FOLDER_ID is not set. "
            "Add it as a GitHub Actions secret or to your .env file."
        )

    from googleapiclient.http import MediaIoBaseDownload

    service = _drive_service()

    query = (
        f"'{GOOGLE_DRIVE_FOLDER_ID}' in parents"
        f" and name contains '{date_prefix}'"
        f" and mimeType != 'application/vnd.google-apps.folder'"
        f" and trashed = false"
    )

    response = (
        service.files()
        .list(
            q=query,
            fields="files(id, name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )

    drive_files = response.get("files", [])
    excel_files = [
        f for f in drive_files if f["name"].lower().endswith((".xlsx", ".xls", ".csv"))
    ]

    if not excel_files:
        return []

    downloaded: list[str] = []
    for file_info in excel_files:
        dest_path = os.path.join(tmp_dir, file_info["name"])
        request = service.files().get_media(
            fileId=file_info["id"],
            supportsAllDrives=True,
        )
        buffer = io.BytesIO()
        downloader = MediaIoBaseDownload(buffer, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        with open(dest_path, "wb") as fh:
            fh.write(buffer.getvalue())
        print(f"  Downloaded: {file_info['name']}")
        downloaded.append(dest_path)

    return downloaded


# ---------------------------------------------------------------------------
# Syncore web session helpers
# ---------------------------------------------------------------------------

async def syncore_login(client: httpx.AsyncClient) -> bool:
    """Log in to Syncore and store session cookies in client.

    Returns True on success, False on failure.
    """
    login_url = f"{SYNCORE_WEB_BASE}/Account/Login"
    try:
        resp = await client.get(login_url)
        resp.raise_for_status()
    except Exception as exc:
        print(f"  [Syncore] Could not reach login page: {exc}")
        return False

    # Extract the anti-forgery token embedded in the login form
    match = re.search(
        r'<input[^>]+name="__RequestVerificationToken"[^>]+value="([^"]+)"',
        resp.text,
    ) or re.search(
        r'name="__RequestVerificationToken"\s+type="hidden"\s+value="([^"]+)"',
        resp.text,
    )
    if not match:
        print("  [Syncore] Could not find CSRF token on login page")
        return False

    resp = await client.post(
        login_url,
        data={
            "Email": SYNCORE_USERNAME,
            "Password": SYNCORE_PASSWORD,
            "__RequestVerificationToken": match.group(1),
        },
    )

    final_url = str(resp.url)
    print(f"  [Syncore] Post-login redirect URL: {final_url}")
    if "/Account/Login" in final_url:
        print("  [Syncore] Login failed — check SYNCORE_USERNAME / SYNCORE_PASSWORD")
        return False
    if any(seg in final_url for seg in ("/Account/Two", "/Account/MFA", "/Account/Send", "/Account/Verify")):
        print(
            "  [Syncore] Login blocked by MFA — use a service account with MFA disabled"
        )
        return False

    print("  [Syncore] Logged in successfully")
    return True


async def add_tracker_entry(
    client: httpx.AsyncClient, job_id: int, description: str
) -> None:
    """POST a tracker entry to a Syncore job via the web UI endpoint."""
    resp = await client.post(
        f"{SYNCORE_WEB_BASE}/Job/AddTrackerEntryAsync",
        data={"JobId": job_id, "TextColor": 1, "Description": description},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    resp.raise_for_status()
    content_type = resp.headers.get("content-type", "")
    if "text/html" in content_type:
        raise RuntimeError(
            "Syncore returned HTML instead of JSON — session may have expired or MFA is required"
        )


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _extract_po(ref1, ref2, ref3=None, ref4=None) -> tuple[str | None, str | None]:
    """Search all four reference fields for a Syncore PO or SRF number.

    Returns (po_number, srf_number); at most one will be non-None.
    Also recognises embedded POs in compound strings like
    "111787-32043-1" → po "32043-1" (prefers 5-digit job prefix over 6-digit).
    """
    for raw in (ref1, ref2, ref3, ref4):
        val = str(raw or "").strip()
        if not val:
            continue
        # Strip common "PO " prefix (e.g. "PO 31987-1" → "31987-1")
        if val.upper().startswith("PO "):
            val = val[3:].strip()
        # Exact PO match
        if PO_PATTERN.match(val):
            return val, None
        # SRF reference number (e.g. "SRF-1234")
        if SRF_PATTERN.match(val):
            return None, val
        # Embedded PO in compound string (e.g. "111787-32043-1" → "32043-1").
        # Split on dashes, check all consecutive-pair candidates; prefer the
        # last 5-digit prefix match over any 6-digit prefix match.
        parts = val.split("-")
        five_digit_match = six_digit_match = None
        for i in range(len(parts) - 1):
            candidate = f"{parts[i]}-{parts[i + 1]}"
            if parts[i].isdigit() and PO_PATTERN.match(candidate):
                if len(parts[i]) == 5:
                    five_digit_match = candidate  # keep updating → last wins
                elif len(parts[i]) == 6:
                    six_digit_match = candidate
        embedded = five_digit_match or six_digit_match
        if embedded:
            return embedded, None
    return None, None


def _file_type(filename: str) -> str:
    """Return 'OB' (OzlinkOB / outbound sales orders) or '3P' (Ozlink3P / third-party POs)."""
    name = filename.lower()
    if "ozlinkob" in name:
        return "OB"
    if "ozlink3p" in name:
        return "3P"
    return "unknown"


def _parse_csv_file(filepath: str, file_type: str = "unknown") -> list[dict]:
    """Parse a UPS CSV file. Returns same format as the Excel parser."""
    import csv

    col_tracking = col_neg_charge = col_ref1 = col_ref2 = col_shipper = None
    col_ref3 = col_ref4 = None  # Shipment Reference Number / Value 2

    with open(filepath, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        rows_raw = list(reader)

    # Find header row
    header_row_idx = None
    for i, row in enumerate(rows_raw):
        normalized = [str(c or "").strip().lower() for c in row]
        if "tracking number" in normalized:
            header_row_idx = i
            col_tracking = normalized.index("tracking number")
            if "neg total charge" in normalized:
                col_neg_charge = normalized.index("neg total charge")
            if "package reference number value 1" in normalized:
                col_ref1 = normalized.index("package reference number value 1")
            if "package reference number value 2" in normalized:
                col_ref2 = normalized.index("package reference number value 2")
            # Name column: OB uses "Ship To Name"; 3P/unknown use "Shipper Name"
            if file_type == "OB":
                if "ship to name" in normalized:
                    col_shipper = normalized.index("ship to name")
            else:
                if "shipper name" in normalized:
                    col_shipper = normalized.index("shipper name")
            # Columns F/G: Shipment Reference Number and Value 2
            for col_idx, hdr in enumerate(normalized):
                if "shipment reference number" in hdr and "value" not in hdr and col_ref3 is None:
                    col_ref3 = col_idx
                elif "shipment reference" in hdr and "value" in hdr and col_ref4 is None:
                    col_ref4 = col_idx
            break

    if header_row_idx is None:
        raise ValueError(f"Header row not found in {os.path.basename(filepath)}")

    missing = [
        name
        for name, idx in [
            ("Tracking Number", col_tracking),
            ("Neg Total Charge", col_neg_charge),
            ("Package Reference Number Value 1", col_ref1),
            ("Package Reference Number Value 2", col_ref2),
        ]
        if idx is None
    ]
    if missing:
        raise ValueError(f"Missing columns in {os.path.basename(filepath)}: {missing}")

    rows: list[dict] = []
    for row in rows_raw[header_row_idx + 1:]:
        if len(row) <= col_tracking:
            continue
        tracking = str(row[col_tracking] or "").strip()
        if not tracking.startswith("1Z"):
            continue

        try:
            ups_cost = float(row[col_neg_charge] or 0)
        except (TypeError, ValueError, IndexError):
            ups_cost = 0.0

        ref1 = row[col_ref1] if col_ref1 < len(row) else None
        ref2 = row[col_ref2] if col_ref2 < len(row) else None
        ref3 = row[col_ref3] if col_ref3 is not None and col_ref3 < len(row) else None
        ref4 = row[col_ref4] if col_ref4 is not None and col_ref4 < len(row) else None
        po_number, srf_number = _extract_po(ref1, ref2, ref3, ref4)
        shipper = str(row[col_shipper] or "").strip() if col_shipper is not None and col_shipper < len(row) else ""
        rows.append({"tracking": tracking, "ups_cost": ups_cost, "po_number": po_number, "srf_number": srf_number, "shipper": shipper})

    return rows


def parse_ups_file(filepath: str, file_type: str = "unknown") -> list[dict]:
    """
    Parse a UPS Excel or CSV file.  Returns a list of dicts with keys:
        tracking  – UPS tracking number (starts with 1Z)
        ups_cost  – Neg Total Charge (float)
        po_number – Syncore PO/SO number e.g. "31987-1" (str or None)
    """
    if filepath.lower().endswith(".csv"):
        return _parse_csv_file(filepath, file_type)

    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    col_tracking = col_neg_charge = col_ref1 = col_ref2 = col_shipper = None
    col_ref3 = col_ref4 = None  # Shipment Reference Number / Value 2
    header_row_idx = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [str(c or "").strip().lower() for c in row]
        if "tracking number" in normalized:
            header_row_idx = row_idx
            col_tracking = normalized.index("tracking number")
            if "neg total charge" in normalized:
                col_neg_charge = normalized.index("neg total charge")
            if "package reference number value 1" in normalized:
                col_ref1 = normalized.index("package reference number value 1")
            if "package reference number value 2" in normalized:
                col_ref2 = normalized.index("package reference number value 2")
            # Name column: OB uses "Ship To Name"; 3P/unknown use "Shipper Name"
            if file_type == "OB":
                if "ship to name" in normalized:
                    col_shipper = normalized.index("ship to name")
            else:
                if "shipper name" in normalized:
                    col_shipper = normalized.index("shipper name")
            # Columns F/G: Shipment Reference Number and Value 2
            for col_idx, hdr in enumerate(normalized):
                if "shipment reference number" in hdr and "value" not in hdr and col_ref3 is None:
                    col_ref3 = col_idx
                elif "shipment reference" in hdr and "value" in hdr and col_ref4 is None:
                    col_ref4 = col_idx
            break

    if header_row_idx is None:
        raise ValueError(f"Header row not found in {os.path.basename(filepath)}")

    missing = [
        name
        for name, idx in [
            ("Tracking Number", col_tracking),
            ("Neg Total Charge", col_neg_charge),
            ("Package Reference Number Value 1", col_ref1),
            ("Package Reference Number Value 2", col_ref2),
        ]
        if idx is None
    ]
    if missing:
        raise ValueError(f"Missing columns in {os.path.basename(filepath)}: {missing}")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        tracking = str(row[col_tracking] or "").strip()
        if not tracking.startswith("1Z"):
            continue

        try:
            ups_cost = float(row[col_neg_charge] or 0)
        except (TypeError, ValueError):
            ups_cost = 0.0

        ref3 = row[col_ref3] if col_ref3 is not None else None
        ref4 = row[col_ref4] if col_ref4 is not None else None
        po_number, srf_number = _extract_po(row[col_ref1], row[col_ref2], ref3, ref4)
        shipper = str(row[col_shipper] or "").strip() if col_shipper is not None else ""
        rows.append({"tracking": tracking, "ups_cost": ups_cost, "po_number": po_number, "srf_number": srf_number, "shipper": shipper})

    wb.close()
    return rows


def group_by_po(rows: list[dict]) -> tuple[dict, list[dict], list[dict]]:
    """
    Group rows by PO number.

    Returns:
        groups   – {po_number: {po_number, job_id, tracking_numbers, total_cost}}
        no_po    – rows where no Syncore PO or SRF number was found
        srf_rows – rows that have an SRF reference number (not a Syncore PO)
    """
    groups: dict[str, dict] = {}
    no_po: list[dict] = []
    srf_rows: list[dict] = []

    for row in rows:
        po = row["po_number"]
        srf = row.get("srf_number")

        if po:
            if po not in groups:
                job_id = int(po.split("-")[0])  # "31987" from "31987-1"
                groups[po] = {
                    "po_number": po,
                    "job_id": job_id,
                    "tracking_numbers": [],
                    "total_cost": 0.0,
                }
            groups[po]["tracking_numbers"].append(row["tracking"])
            groups[po]["total_cost"] = round(groups[po]["total_cost"] + row["ups_cost"], 2)
        elif srf:
            srf_rows.append(row)
        else:
            no_po.append(row)

    return groups, no_po, srf_rows


# ---------------------------------------------------------------------------
# Job Log entry formatting
# ---------------------------------------------------------------------------

def build_log_entry(po_number: str, tracking_numbers: list[str], total_cost: float, label: str = "PO") -> str:
    today_str = datetime.now(ZoneInfo("America/Los_Angeles")).strftime("%m/%d/%Y")
    pkg_count = len(tracking_numbers)
    pkg_label = "package" if pkg_count == 1 else "packages"
    tracking_str = "\n  ".join(tracking_numbers)

    return (
        f"UPS Tracking Import — {today_str}\n"
        f"{label}: {po_number}\n"
        f"Packages: {pkg_count} {pkg_label}\n"
        f"Tracking Number(s):\n  {tracking_str}\n"
        f"UPS Shipping Cost: ${total_cost:.2f}"
    )


# ---------------------------------------------------------------------------
# Email notifications
# ---------------------------------------------------------------------------

def send_email(subject: str, body: str, to: str | None = None) -> None:
    """Send a plain-text notification email via Gmail SMTP.

    ``to`` defaults to CSR_EMAIL when omitted.
    """
    recipient = to or CSR_EMAIL
    if not all([GMAIL_USER, GMAIL_APP_PASSWORD, recipient]):
        print(f"[EMAIL NOT CONFIGURED]\nSubject: {subject}\n{body}\n")
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = GMAIL_USER
    msg["To"] = recipient
    msg.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, recipient, msg.as_string())
        print(f"  ✉  Email sent to {recipient}: {subject}")
    except Exception as exc:
        print(f"  ✗  Email failed: {exc}")


# ---------------------------------------------------------------------------
# Per-file processing
# ---------------------------------------------------------------------------

async def process_file(filepath: str, client: httpx.AsyncClient) -> dict:
    """
    Parse one UPS Excel/CSV file and write tracker entries to Syncore.

    Returns a summary dict with keys:
        file, file_type, logs_added, logs_added_details,
        manual_entries, errors, skipped_no_po, skipped_srf
    """
    filename = os.path.basename(filepath)
    ftype = _file_type(filename)
    label = "SO" if ftype == "OB" else "PO"
    print(f"\n  [{filename}]  ({ftype})")

    result: dict = {
        "file": filename,
        "file_type": ftype,
        "logs_added": 0,
        "logs_added_details": [],
        "manual_entries": [],
        "errors": [],
        "skipped_no_po": [],
        "skipped_srf": [],
    }

    try:
        rows = parse_ups_file(filepath, ftype)
    except Exception as exc:
        result["errors"].append(f"Parse error: {exc}")
        return result

    print(f"    Rows found: {len(rows)}")

    groups, no_po, srf_rows = group_by_po(rows)

    if no_po:
        print(f"    Skipped (no {label}#): {len(no_po)}")
        result["skipped_no_po"] = no_po

    if srf_rows:
        print(f"    Skipped (SRF ref): {len(srf_rows)}")
        result["skipped_srf"] = srf_rows

    for po_number, group in sorted(groups.items()):
        job_id = group["job_id"]
        tracking_numbers = group["tracking_numbers"]
        total_cost = group["total_cost"]

        print(
            f"    {label} {po_number} → Job #{job_id} | "
            f"{len(tracking_numbers)} pkg(s) | ${total_cost:.2f}"
        )

        log_text = build_log_entry(po_number, tracking_numbers, total_cost, label)

        try:
            await add_tracker_entry(client, job_id, log_text)
            print(f"      ✓ Tracker entry added")
            result["logs_added"] += 1
            result["logs_added_details"].append({
                "po_number": po_number,
                "job_id": job_id,
                "tracking_numbers": tracking_numbers,
                "total_cost": total_cost,
            })
        except httpx.HTTPStatusError as exc:
            detail = exc.response.text[:200].strip()
            msg = f"HTTP {exc.response.status_code} for Job #{job_id} ({label} {po_number})"
            print(f"      ✗ {msg} — added to manual entry email")
            if detail:
                print(f"        Syncore says: {detail}")
            result["manual_entries"].append({
                "job_id": job_id,
                "po_number": po_number,
                "log_text": log_text,
            })
        except Exception as exc:
            msg = f"Unexpected error for Job #{job_id} ({label} {po_number}): {exc}"
            print(f"      ✗ {msg}")
            result["errors"].append(msg)

    return result


# ---------------------------------------------------------------------------
# Run summary email
# ---------------------------------------------------------------------------

def _render_file_section(results: list[dict], label: str, name_col_label: str) -> list[str]:
    """Render added/manual/no-match lines for one group of result dicts (OB or 3P)."""
    sep = "─" * 48
    lines: list[str] = []

    added = [d for r in results for d in r["logs_added_details"]]
    manual = [e for r in results for e in r["manual_entries"]]
    no_match = [t for r in results for t in r["skipped_no_po"]]

    # ── Successfully added ────────────────────────────────────────────────
    lines.append(f"✓ ADDED TO SYNCORE ({len(added)} job log {'entry' if len(added) == 1 else 'entries'})")
    lines.append(sep)
    if added:
        for item in added:
            pkg_label = "package" if len(item["tracking_numbers"]) == 1 else "packages"
            lines.append(
                f"{label} {item['po_number']} — Job #{item['job_id']}\n"
                f"  {len(item['tracking_numbers'])} {pkg_label} | ${item['total_cost']:.2f}\n"
                + "\n".join(f"  {t}" for t in item["tracking_numbers"])
            )
    else:
        lines.append("  (none)")
    lines.append("")

    # ── Manual entry required ─────────────────────────────────────────────
    if manual:
        lines.append(f"⚠ MANUAL ENTRY REQUIRED ({len(manual)} {'entry' if len(manual) == 1 else 'entries'})")
        lines.append(sep)
        lines.append("These could not be posted to Syncore automatically.")
        lines.append("Open each job and paste the text below into the Job Log tab.\n")
        for entry in manual:
            lines.append(f"Job #{entry['job_id']} — {label} {entry['po_number']}")
            lines.append(entry["log_text"])
        lines.append("")

    # ── No match found ────────────────────────────────────────────────────
    if no_match:
        n = len(no_match)
        lines.append(f"⚠ NO {label} NUMBER FOUND ({n} tracking {'number' if n == 1 else 'numbers'})")
        lines.append(sep)
        for row in no_match:
            name_val = row.get("shipper", "")
            name_line = f"  {name_col_label}: {name_val}" if name_val else ""
            lines.append(
                f"  {row['tracking']}  |  ${row['ups_cost']:.2f}"
                + (f"\n{name_line}" if name_line else "")
            )
        lines.append("")

    return lines


def _build_srf_email(today_str: str, srf_rows: list[dict]) -> str:
    """Build the body for the SRF tracking email sent to accounting."""
    sep = "─" * 48
    lines = [
        "SRF Tracking Information",
        "",
        sep,
    ]
    for row in srf_rows:
        shipper_val = row.get("shipper", "")
        shipper_line = f"  Shipper: {shipper_val}" if shipper_val else ""
        lines.append(
            f"  SRF: {row.get('srf_number', '')}  |  {row['tracking']}  |  ${row['ups_cost']:.2f}"
            + (f"\n{shipper_line}" if shipper_line else "")
        )
    return "\n".join(lines)


def _build_run_summary(
    today_str: str,
    ob_results: list[dict],
    p3_results: list[dict],
    suboutcorex_names: list[str],
    errors: list[str],
) -> str:
    eq = "═" * 52
    all_files = [r["file"] for r in ob_results + p3_results]
    lines: list[str] = [
        f"UPS Tracking Import — {today_str}",
        f"Files processed: {len(all_files)}"
        + (f"  ({', '.join(all_files)})" if all_files else ""),
        "",
    ]

    # ── Outbound (OzlinkOB) ───────────────────────────────────────────────
    lines.append(f"{eq}")
    lines.append("  OUTBOUND — OzlinkOB (Sales Orders)")
    lines.append(f"{eq}")
    if ob_results:
        lines.extend(_render_file_section(ob_results, label="SO", name_col_label="Ship To"))
    else:
        lines.append("  (no OzlinkOB file processed today)")
        lines.append("")

    # ── 3rd Party (Ozlink3P) ──────────────────────────────────────────────
    lines.append(f"{eq}")
    lines.append("  3RD PARTY — Ozlink3P (Purchase Orders)")
    lines.append(f"{eq}")
    if p3_results:
        lines.extend(_render_file_section(p3_results, label="PO", name_col_label="Shipper"))
    else:
        lines.append("  (no Ozlink3P file processed today)")
        lines.append("")

    # ── Suboutcorex files ─────────────────────────────────────────────────
    sep = "─" * 48
    if suboutcorex_names:
        lines.append(f"⚠ SKIPPED — SUBOUTCOREX FILES ({len(suboutcorex_names)} file{'s' if len(suboutcorex_names) != 1 else ''})")
        lines.append(sep)
        lines.append("These files require manual review and were not imported:")
        lines.extend(f"  • {n}" for n in suboutcorex_names)
        lines.append("")

    # ── Errors ────────────────────────────────────────────────────────────
    if errors:
        lines.append(f"⚠ ERRORS ({len(errors)})")
        lines.append(sep)
        lines.extend(f"  • {e}" for e in errors)
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main() -> None:
    today_str = datetime.now(ZoneInfo("America/Los_Angeles")).strftime("%Y-%m-%d")
    print(f"{'='*55}")
    print(f"  UPS Tracking Import  —  {today_str}")
    print(f"{'='*55}")

    suboutcorex_names: list[str] = []

    if not SYNCORE_USERNAME or not SYNCORE_PASSWORD:
        print("\nERROR: SYNCORE_USERNAME and/or SYNCORE_PASSWORD are not set.")
        return

    # Log in to Syncore once; reuse the session for all jobs
    async with httpx.AsyncClient(
        timeout=HTTP_TIMEOUT, follow_redirects=True
    ) as syncore_client:
        logged_in = await syncore_login(syncore_client)
        if not logged_in:
            send_email(
                subject=f"UPS Import Failed — Syncore Login Error ({today_str})",
                body=(
                    "The UPS import could not log in to Syncore.\n\n"
                    "Check that SYNCORE_USERNAME and SYNCORE_PASSWORD are correct."
                ),
            )
            return

        # Download today's files to a temp directory
        with tempfile.TemporaryDirectory() as tmp_dir:
            try:
                all_files = download_todays_files(tmp_dir)
            except Exception as exc:
                print(f"\nERROR downloading files from Google Drive: {exc}")
                send_email(
                    subject=f"UPS Import Failed — Drive Error ({today_str})",
                    body=f"The UPS import could not download files from Google Drive:\n\n{exc}",
                )
                return

            if not all_files:
                print(f"\n  No UPS files found for {date.today().strftime('%Y%m%d')}. Nothing to do.")
                return

            print(f"  Files found: {len(all_files)}\n")

            # Split out suboutcorex files
            suboutcorex = [f for f in all_files if "suboutcorex" in os.path.basename(f).lower()]
            processable = [f for f in all_files if "suboutcorex" not in os.path.basename(f).lower()]

            if suboutcorex:
                suboutcorex_names = [os.path.basename(f) for f in suboutcorex]
                names = suboutcorex_names
                print(f"  [NOTICE] Skipping {len(suboutcorex)} suboutcorex file(s): {', '.join(names)}")
                send_email(
                    subject=f"UPS Import — Manual Review Required ({today_str})",
                    body=(
                        f"The following UPS file(s) contain 'suboutcorex' in the filename "
                        f"and were NOT automatically imported. Please review them manually:\n\n"
                        + "\n".join(f"  • {n}" for n in names)
                    ),
                )

            if not processable:
                print("  No processable files remaining.")
                return

            all_results = []
            for filepath in processable:
                result = await process_file(filepath, syncore_client)
                all_results.append(result)

    # async with syncore_client and temp directory are both cleaned up here

    ob_results = [r for r in all_results if r["file_type"] == "OB"]
    p3_results = [r for r in all_results if r["file_type"] != "OB"]

    total_logs = sum(r["logs_added"] for r in all_results)
    all_manual = [e for r in all_results for e in r["manual_entries"]]
    all_errors = [err for r in all_results for err in r["errors"]]
    all_skipped_no_po = [t for r in all_results for t in r["skipped_no_po"]]
    # SRF rows only come from 3P files
    all_skipped_srf = [t for r in p3_results for t in r["skipped_srf"]]

    print(f"\n{'='*55}")
    print(f"  Job Logs added   : {total_logs}")
    print(f"  Manual entries   : {len(all_manual)}")
    print(f"  SRF rows         : {len(all_skipped_srf)}")
    print(f"  Errors           : {len(all_errors)}")
    print(f"{'='*55}\n")

    # Send SRF info to accounting
    if all_skipped_srf:
        srf_date = datetime.now(ZoneInfo("America/Los_Angeles")).strftime("%m/%d/%Y")
        send_email(
            subject=f"SRF Tracking Info ({srf_date})",
            body=_build_srf_email(today_str, all_skipped_srf),
            to=ACCOUNTING_EMAIL,
        )

    if all_manual:
        sep = "-" * 48
        blocks = []
        for entry in all_manual:
            # Determine label from log_text (starts with "SO:" or "PO:")
            first_line = entry["log_text"].split("\n")[1] if "\n" in entry["log_text"] else ""
            ref_label = "SO" if first_line.startswith("SO:") else "PO"
            blocks.append(
                f"Job #{entry['job_id']} — {ref_label} {entry['po_number']}\n"
                f"{sep}\n"
                f"{entry['log_text']}"
            )
        send_email(
            subject=f"UPS Import — Manual Entry Required ({today_str})",
            body=(
                f"The UPS tracking import ran on {today_str}.\n"
                f"The Syncore job log API is currently unavailable, so the "
                f"{len(all_manual)} entry/entries below could not be added automatically.\n\n"
                f"For each job, open it in Syncore and paste the entry into the Job Log tab.\n\n"
                f"{'=' * 48}\n"
                + f"\n\n{'=' * 48}\n".join(blocks)
                + f"\n{'=' * 48}"
            ),
        )

    if all_errors:
        send_email(
            subject=f"UPS Import Errors ({today_str})",
            body=(
                f"The UPS tracking import on {today_str} encountered unexpected errors:\n\n"
                + "\n".join(f"  • {err}" for err in all_errors)
            ),
        )

    # Always send the CSR a full run summary
    needs_attention = bool(all_manual or all_skipped_no_po or suboutcorex_names or all_errors)
    summary_subject = (
        f"UPS Import — Action Required ({today_str})"
        if needs_attention
        else f"UPS Import Complete ({today_str})"
    )
    send_email(
        subject=summary_subject,
        body=_build_run_summary(
            today_str,
            ob_results,
            p3_results,
            suboutcorex_names,
            all_errors,
        ),
    )


if __name__ == "__main__":
    asyncio.run(main())
